"""
AlphaAlign -- Excel Survey Importer (v2).

Parses completed Excel survey files in both formats:
  - v2: Question-per-row "Pairwise Survey" sheet with dropdown answers
  - v1 (legacy): Raw NxN "Pairwise Comparison" matrix

Extracts respondent data, pairwise comparison values, alternative scores,
strategic importance, and comments.
"""

from typing import Dict, List, Optional
from openpyxl import load_workbook
import io


# ── Dropdown label to Saaty value mapping ──
DROPDOWN_TO_VALUE = {
    "A is overwhelmingly more important": 9.0,
    "A is much more important": 7.0,
    "A is clearly more important": 5.0,
    "A is slightly more important": 3.0,
    "Equal importance": 1.0,
    "B is slightly more important": 1/3,
    "B is clearly more important": 1/5,
    "B is much more important": 1/7,
    "B is overwhelmingly more important": 1/9,
}

# ── Outcome score label to numeric value ──
OUTCOME_LABEL_TO_VALUE = {
    "1 - Very Poor Fit": 1.0,
    "2 - Poor Fit": 2.0,
    "3 - Moderate Fit": 3.0,
    "4 - Good Fit": 4.0,
    "5 - Excellent Fit": 5.0,
}


def import_survey(file_content: bytes, criteria_map: Dict[str, str]) -> Dict:
    """
    Import a completed survey file (v1 or v2 format).

    Args:
        file_content: Raw bytes of the Excel file.
        criteria_map: {criterion_name: criterion_id} mapping.

    Returns:
        Dict with:
            respondent: {name, role, email}
            pairwise_values: [{criterion_a_id, criterion_b_id, value}]
            alternative_scores: [{criterion_id, outcome, score}]
            strategic_importance: [{item_name, type, value, justification, ...}]
            comments: str
            metadata: {project_id, ...}
    """
    wb = load_workbook(io.BytesIO(file_content), data_only=True, keep_vba=True)

    respondent = _parse_respondent(wb)
    pairwise = _parse_pairwise(wb, criteria_map)
    alt_scores = _parse_alternative_scores(wb, criteria_map)
    strategic = _parse_strategic_importance(wb)
    comments = _parse_comments(wb)
    metadata = _parse_metadata(wb)

    return {
        "respondent": respondent,
        "pairwise_values": pairwise,
        "alternative_scores": alt_scores,
        "strategic_importance": strategic,
        "comments": comments,
        "metadata": metadata,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  Respondent
# ══════════════════════════════════════════════════════════════════════════════

def _parse_respondent(wb) -> Dict:
    """Parse respondent info from Welcome or Respondent Profile sheet."""
    # Try v2 format (Welcome sheet)
    if "Welcome" in wb.sheetnames:
        ws = wb["Welcome"]
        data = {}
        field_map = {"Name": "name", "Role": "role", "Email": "email"}
        # In v2, respondent info starts after the progress section
        for row in ws.iter_rows(min_row=1, max_row=50, max_col=2, values_only=False):
            if len(row) >= 2 and row[0].value and str(row[0].value).strip() in field_map:
                key = field_map[str(row[0].value).strip()]
                val = row[1].value or ""
                # Skip placeholder text
                val_str = str(val).strip()
                if val_str and not val_str.startswith("Enter your") and not val_str.startswith("e.g.") and val_str != "(optional)":
                    data[key] = val_str
        return {
            "name": data.get("name", ""),
            "role": data.get("role", ""),
            "email": data.get("email", ""),
        }

    # Try v1 format (Respondent Profile sheet)
    if "Respondent Profile" in wb.sheetnames:
        ws = wb["Respondent Profile"]
        data = {}
        field_map = {"Name": "name", "Role": "role", "Email": "email"}
        for row in ws.iter_rows(min_row=3, max_row=10, max_col=2, values_only=False):
            if len(row) >= 2 and row[0].value in field_map:
                key = field_map[row[0].value]
                val = row[1].value or ""
                data[key] = str(val).strip()
        return {
            "name": data.get("name", ""),
            "role": data.get("role", ""),
            "email": data.get("email", ""),
        }

    return {"name": "", "role": "", "email": ""}


# ══════════════════════════════════════════════════════════════════════════════
#  Pairwise Comparisons
# ══════════════════════════════════════════════════════════════════════════════

def _parse_pairwise(wb, criteria_map: Dict[str, str]) -> List[Dict]:
    """Parse pairwise comparisons from v2 or v1 format."""
    # Try v2 format first (question-per-row)
    if "Pairwise Survey" in wb.sheetnames:
        return _parse_pairwise_v2(wb, criteria_map)

    # Fall back to v1 format (matrix)
    if "Pairwise Comparison" in wb.sheetnames:
        return _parse_pairwise_v1(wb, criteria_map)

    return []


def _parse_pairwise_v2(wb, criteria_map: Dict[str, str]) -> List[Dict]:
    """Parse v2 question-per-row format."""
    ws = wb["Pairwise Survey"]
    results = []

    for row in range(5, ws.max_row + 1):
        # Column B: Criterion A name, Column D: Criterion B name
        crit_a_name = ws.cell(row=row, column=2).value
        crit_b_name = ws.cell(row=row, column=4).value

        if not crit_a_name or not crit_b_name:
            continue

        crit_a_name = str(crit_a_name).strip()
        crit_b_name = str(crit_b_name).strip()

        # Look up criterion IDs
        crit_a_id = criteria_map.get(crit_a_name)
        crit_b_id = criteria_map.get(crit_b_name)
        if not crit_a_id or not crit_b_id:
            # Try hidden ID columns (G and H)
            crit_a_id = crit_a_id or str(ws.cell(row=row, column=7).value or "").strip()
            crit_b_id = crit_b_id or str(ws.cell(row=row, column=8).value or "").strip()
        if not crit_a_id or not crit_b_id:
            continue

        # Try to get the computed Saaty value from hidden column F
        saaty_value = ws.cell(row=row, column=6).value
        if saaty_value is not None:
            try:
                value = float(saaty_value)
                if value > 0:
                    results.append({
                        "criterion_a_id": crit_a_id,
                        "criterion_b_id": crit_b_id,
                        "value": value,
                    })
                    continue
            except (ValueError, TypeError):
                pass

        # Fallback: convert dropdown label to value directly
        dropdown_text = ws.cell(row=row, column=5).value
        if dropdown_text:
            dropdown_text = str(dropdown_text).strip()
            value = DROPDOWN_TO_VALUE.get(dropdown_text)
            if value is not None and value > 0:
                results.append({
                    "criterion_a_id": crit_a_id,
                    "criterion_b_id": crit_b_id,
                    "value": value,
                })

    return results


def _parse_pairwise_v1(wb, criteria_map: Dict[str, str]) -> List[Dict]:
    """Parse v1 matrix format (legacy backward compatibility)."""
    ws = wb["Pairwise Comparison"]
    results = []
    header_row = 4
    criteria_cols = {}

    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and str(val).strip() in criteria_map:
            criteria_cols[col] = str(val).strip()

    criteria_rows = {}
    for row in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and str(val).strip() in criteria_map:
            criteria_rows[row] = str(val).strip()

    for row_idx, row_name in criteria_rows.items():
        for col_idx, col_name in criteria_cols.items():
            if row_name == col_name:
                continue
            row_order = list(criteria_rows.values())
            col_order = list(criteria_cols.values())
            if row_order.index(row_name) >= col_order.index(col_name):
                continue

            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                try:
                    value = _parse_ahp_value(cell_value)
                    if value is not None and value > 0:
                        results.append({
                            "criterion_a_id": criteria_map[row_name],
                            "criterion_b_id": criteria_map[col_name],
                            "value": value,
                        })
                except (ValueError, ZeroDivisionError):
                    pass

    return results


def _parse_ahp_value(value) -> Optional[float]:
    """Parse an AHP value which could be numeric or a fraction string."""
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.strip()
        if "/" in value:
            parts = value.split("/")
            if len(parts) == 2:
                return float(parts[0]) / float(parts[1])
        try:
            return float(value)
        except ValueError:
            return None
    return None


# ══════════════════════════════════════════════════════════════════════════════
#  Alternative Scores (No longer in survey, using DB defaults instead)
# ══════════════════════════════════════════════════════════════════════════════

def _parse_alternative_scores(wb, criteria_map: Dict[str, str]) -> List[Dict]:
    return []


# ══════════════════════════════════════════════════════════════════════════════
#  Strategic Importance (Removed to simplify survey)
# ══════════════════════════════════════════════════════════════════════════════

def _parse_strategic_importance(wb) -> List[Dict]:
    return []


# ══════════════════════════════════════════════════════════════════════════════
#  Comments & Metadata
# ══════════════════════════════════════════════════════════════════════════════

def _parse_comments(wb) -> str:
    if "Comments" not in wb.sheetnames:
        return ""
    ws = wb["Comments"]
    comments = ws.cell(row=5, column=1).value
    return str(comments).strip() if comments else ""


def _parse_metadata(wb) -> Dict:
    if "Metadata" not in wb.sheetnames:
        return {}

    ws = wb["Metadata"]
    data = {}
    for row in range(3, 15):
        key = ws.cell(row=row, column=1).value
        val = ws.cell(row=row, column=2).value
        if key:
            data[str(key).strip()] = str(val).strip() if val else ""

    return data
