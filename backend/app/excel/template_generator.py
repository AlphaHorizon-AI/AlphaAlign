"""
AlphaAlign -- Smart Excel Survey Template Generator (v2).

Redesigned to be user-friendly for C-level respondents:
  - Question-per-row pairwise comparisons (no raw matrix)
  - Plain-language dropdown scale (no reciprocals, no Saaty numbers)
  - Hidden formula columns auto-compute Saaty values
  - Hidden Matrix sheet auto-populates the full NxN comparison matrix
  - Live progress tracking via formulas
  - Conditional formatting for completion
  - Optional VBA macros for interactivity
"""

import io
import os
import tempfile
from datetime import datetime
from typing import List, Dict, Tuple, Optional
from itertools import combinations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule


# ── Style Constants ──────────────────────────────────────────────────────────

BRAND_NAVY = "0D1B3E"
BRAND_BLUE = "1565C0"
BRAND_LIGHT = "E3F2FD"
BRAND_ACCENT = "2196F3"

TITLE_FONT = Font(name="Calibri", bold=True, size=18, color="FFFFFF")
SECTION_FONT = Font(name="Calibri", bold=True, size=13, color=BRAND_NAVY)
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
LABEL_FONT = Font(name="Calibri", bold=True, size=11)
BODY_FONT = Font(name="Calibri", size=11)
HINT_FONT = Font(name="Calibri", size=10, italic=True, color="666666")
SMALL_FONT = Font(name="Calibri", size=9, color="999999")

HEADER_FILL = PatternFill(start_color=BRAND_NAVY, end_color=BRAND_NAVY, fill_type="solid")
BLUE_FILL = PatternFill(start_color=BRAND_BLUE, end_color=BRAND_BLUE, fill_type="solid")
LIGHT_FILL = PatternFill(start_color=BRAND_LIGHT, end_color=BRAND_LIGHT, fill_type="solid")
ALT_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
PROGRESS_FILL = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
DIAGONAL_FILL = PatternFill(start_color="CFD8DC", end_color="CFD8DC", fill_type="solid")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
MEDIUM_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="medium", color="999999"),
)

# ── Scale Configuration ──────────────────────────────────────────────────────

SCALE_OPTIONS = [
    ("A is overwhelmingly more important", 9),
    ("A is much more important", 7),
    ("A is clearly more important", 5),
    ("A is slightly more important", 3),
    ("Equal importance", 1),
    ("B is slightly more important", 1/3),
    ("B is clearly more important", 1/5),
    ("B is much more important", 1/7),
    ("B is overwhelmingly more important", 1/9),
]

SCALE_LABELS = [label for label, _ in SCALE_OPTIONS]
SCALE_VALUES = {label: value for label, value in SCALE_OPTIONS}

OUTCOME_SCALE = [
    "1 - Very Poor Fit",
    "2 - Poor Fit",
    "3 - Moderate Fit",
    "4 - Good Fit",
    "5 - Excellent Fit",
]


# ── VBA Code ─────────────────────────────────────────────────────────────────

VBA_CODE = '''
Sub CheckProgress()
    Dim wsS As Worksheet, wsW As Worksheet
    Dim lastRow As Long, answered As Long, total As Long, i As Long
    
    Set wsS = ThisWorkbook.Worksheets("Pairwise Survey")
    Set wsW = ThisWorkbook.Worksheets("Welcome")
    
    lastRow = wsS.Cells(wsS.Rows.Count, "B").End(xlUp).Row
    total = 0
    answered = 0
    
    For i = 5 To lastRow
        If wsS.Cells(i, 2).Value <> "" Then
            total = total + 1
            If wsS.Cells(i, 5).Value <> "" Then
                answered = answered + 1
            End If
        End If
    Next i
    
    MsgBox "Pairwise Comparisons: " & answered & " of " & total & " completed (" & _
           Format(IIf(total > 0, answered / total, 0), "0%") & ")" & vbCrLf & vbCrLf & _
           IIf(answered = total, "All comparisons complete!", _
           "You have " & (total - answered) & " comparisons remaining."), _
           IIf(answered = total, vbInformation, vbExclamation), "AlphaAlign Progress"
End Sub

Sub GoToNextEmpty()
    Dim wsS As Worksheet
    Dim lastRow As Long, i As Long
    
    Set wsS = ThisWorkbook.Worksheets("Pairwise Survey")
    lastRow = wsS.Cells(wsS.Rows.Count, "B").End(xlUp).Row
    
    For i = 5 To lastRow
        If wsS.Cells(i, 2).Value <> "" And wsS.Cells(i, 5).Value = "" Then
            wsS.Activate
            wsS.Cells(i, 5).Select
            Exit Sub
        End If
    Next i
    
    MsgBox "All pairwise comparisons are filled in!", vbInformation, "AlphaAlign"
End Sub

Sub HighlightMissing()
    Dim wsS As Worksheet
    Dim lastRow As Long, i As Long, count As Long
    
    Set wsS = ThisWorkbook.Worksheets("Pairwise Survey")
    lastRow = wsS.Cells(wsS.Rows.Count, "B").End(xlUp).Row
    count = 0
    
    For i = 5 To lastRow
        If wsS.Cells(i, 2).Value <> "" Then
            If wsS.Cells(i, 5).Value = "" Then
                wsS.Range("A" & i & ":E" & i).Interior.Color = RGB(255, 235, 238)
                count = count + 1
            Else
                wsS.Range("A" & i & ":E" & i).Interior.Color = RGB(232, 245, 233)
            End If
        End If
    Next i
    
    If count = 0 Then
        MsgBox "All comparisons are complete!", vbInformation, "AlphaAlign"
    Else
        MsgBox count & " comparisons still need your input." & vbCrLf & _
               "They have been highlighted in red.", vbExclamation, "AlphaAlign"
    End If
End Sub
'''


# ══════════════════════════════════════════════════════════════════════════════
#  Main Entry Point
# ══════════════════════════════════════════════════════════════════════════════

def generate_survey_template(
    project_id: str,
    project_name: str,
    company_name: str,
    criteria: List[Dict],
    include_strategic_importance: bool = False,
) -> io.BytesIO:
    """
    Generate a smart Excel survey template.

    The survey is intentionally simple — respondents ONLY fill in pairwise
    comparisons (plain-language dropdowns) and optional comments.
    Outcome scoring is handled by expert-derived defaults on the admin side.

    Returns a BytesIO buffer containing the .xlsm or .xlsx file.
    """
    wb = Workbook()

    # Generate all unique pairs
    n = len(criteria)
    pairs = list(combinations(range(n), 2))  # (i, j) where i < j
    total_pairs = len(pairs)

    # Build sheets — only 3 user-facing + hidden helper sheets
    _create_config_sheet(wb, criteria)
    _create_welcome_sheet(wb, project_name, company_name, criteria, total_pairs)
    _create_pairwise_survey_sheet(wb, criteria, pairs)
    _create_comments_sheet(wb)
    _create_matrix_sheet(wb, criteria, pairs)
    _create_metadata_sheet(wb, project_id, project_name, company_name, criteria)

    # Remove default "Sheet"
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save as .xlsx first
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Try to inject VBA macros (Windows + Excel installed)
    xlsm_buffer = _try_inject_vba(buffer)
    if xlsm_buffer:
        return xlsm_buffer

    # Fallback: return .xlsx (still fully smart with formulas)
    return buffer


def get_file_extension(buffer: io.BytesIO) -> str:
    """Check if the buffer is an xlsm or xlsx by peeking at the content."""
    # We track this via a simple heuristic -- xlsm files have vbaProject.bin
    import zipfile
    try:
        buffer.seek(0)
        with zipfile.ZipFile(buffer, 'r') as zf:
            if 'xl/vbaProject.bin' in zf.namelist():
                return ".xlsm"
    except Exception:
        pass
    finally:
        buffer.seek(0)
    return ".xlsx"


# ══════════════════════════════════════════════════════════════════════════════
#  Config Sheet (Very Hidden)
# ══════════════════════════════════════════════════════════════════════════════

def _create_config_sheet(wb: Workbook, criteria: List[Dict]):
    """Hidden lookup tables for dropdowns and formulas."""
    ws = wb.create_sheet("Config")
    ws.sheet_state = "veryHidden"

    # ── Saaty Scale Lookup Table ──
    ws.cell(row=1, column=1, value="Label")
    ws.cell(row=1, column=2, value="Value")
    for i, (label, value) in enumerate(SCALE_OPTIONS, start=2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=round(value, 6))

    # ── Outcome Score Lookup ──
    ws.cell(row=1, column=4, value="ScoreLabel")
    ws.cell(row=1, column=5, value="ScoreValue")
    for i, label in enumerate(OUTCOME_SCALE, start=2):
        ws.cell(row=i, column=4, value=label)
        ws.cell(row=i, column=5, value=i - 1)  # 1, 2, 3, 4, 5

    # ── Criteria List ──
    ws.cell(row=1, column=7, value="CriterionID")
    ws.cell(row=1, column=8, value="CriterionName")
    for i, c in enumerate(criteria, start=2):
        ws.cell(row=i, column=7, value=c["id"])
        ws.cell(row=i, column=8, value=c["name"])


# ══════════════════════════════════════════════════════════════════════════════
#  Welcome Sheet
# ══════════════════════════════════════════════════════════════════════════════

def _create_welcome_sheet(wb: Workbook, project_name: str, company_name: str,
                          criteria: List[Dict], total_pairs: int):
    ws = wb.create_sheet("Welcome")
    ws.sheet_properties.tabColor = BRAND_NAVY

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 25

    # ── Title Banner ──
    ws.merge_cells("A1:C1")
    title_cell = ws.cell(row=1, column=1, value="AlphaAlign")
    title_cell.font = TITLE_FONT
    title_cell.fill = PatternFill(start_color=BRAND_NAVY, end_color=BRAND_NAVY, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 50

    ws.merge_cells("A2:C2")
    sub = ws.cell(row=2, column=1, value="Strategic AI Architecture Assessment")
    sub.font = Font(name="Calibri", size=12, color=BRAND_BLUE)
    sub.alignment = Alignment(horizontal="center")

    # ── Project Info ──
    row = 4
    for label, value in [
        ("Assessment", project_name),
        ("Company", company_name),
        ("Generated", datetime.now().strftime("%B %d, %Y at %H:%M")),
        ("Comparisons Required", str(total_pairs)),
    ]:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = BODY_FONT
        row += 1

    # ── Live Progress ──
    row += 1
    ws.cell(row=row, column=1, value="YOUR PROGRESS").font = SECTION_FONT
    row += 1

    # Pairwise progress
    n_criteria = len(criteria)
    pw_last_row = 4 + total_pairs  # data starts at row 5 in survey sheet
    ws.cell(row=row, column=1, value="Pairwise Comparisons").font = LABEL_FONT
    ws.cell(row=row, column=2).font = BODY_FONT
    ws.cell(row=row, column=2, value=f'=COUNTA(\'Pairwise Survey\'!E5:E{pw_last_row}) & " of {total_pairs} completed"')
    # Visual progress bar
    ws.cell(row=row, column=3).font = Font(name="Consolas", size=11, color="1B5E20")
    ws.cell(row=row, column=3, value=(
        f'=REPT(UNICHAR(9608),ROUND(COUNTA(\'Pairwise Survey\'!E5:E{pw_last_row})/{total_pairs}*20,0))'
        f'&REPT(UNICHAR(9617),20-ROUND(COUNTA(\'Pairwise Survey\'!E5:E{pw_last_row})/{total_pairs}*20,0))'
    ))
    row += 2

    # ── Respondent Info ──
    ws.cell(row=row, column=1, value="YOUR INFORMATION").font = SECTION_FONT
    row += 1
    for label, placeholder in [
        ("Name", "Enter your full name"),
        ("Role", "e.g. CEO, CFO, CIO, CTO, CISO, COO"),
        ("Email", "your.email@company.com"),
        ("Department", "(optional)"),
    ]:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).fill = LIGHT_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        cell = ws.cell(row=row, column=2)
        cell.border = THIN_BORDER
        cell.fill = INPUT_FILL
        cell.font = HINT_FONT
        # Don't put placeholder as value -- leave empty for user
        row += 1
    row += 1

    # ── Instructions ──
    ws.cell(row=row, column=1, value="HOW TO COMPLETE THIS SURVEY").font = SECTION_FONT
    row += 1

    instructions = [
        ("Step 1:", "Fill in your name, role, and email above."),
        ("Step 2:", "Go to the 'Pairwise Survey' tab. For each pair of criteria, pick your judgement from the dropdown."),
        ("Step 3:", "(Optional) Go to the 'Comments' tab to add any observations."),
        ("Step 4:", "Save the file and return it to the assessment coordinator."),
        ("", ""),
        ("That's it!", "You only need to answer the dropdown questions — no numbers, no math, no technical knowledge required."),
        ("", "The progress bar above updates automatically as you go."),
        ("", ""),
        ("Tip:", 'If unsure about a comparison, just pick "Equal importance" — you can always revisit later.'),
    ]
    for label, desc in instructions:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=desc).font = BODY_FONT
        ws.cell(row=row, column=2).alignment = LEFT_WRAP
        row += 1
    row += 1

    # ── Criteria Reference ──
    ws.cell(row=row, column=1, value="CRITERIA BEING EVALUATED").font = SECTION_FONT
    row += 1
    for i, c in enumerate(criteria, 1):
        ws.cell(row=row, column=1, value=f"{i}. {c['name']}").font = LABEL_FONT
        ws.cell(row=row, column=2, value=c.get("description", "")).font = BODY_FONT
        ws.cell(row=row, column=2).alignment = LEFT_WRAP
        row += 1


# ══════════════════════════════════════════════════════════════════════════════
#  Pairwise Survey Sheet (The Main Innovation)
# ══════════════════════════════════════════════════════════════════════════════

def _create_pairwise_survey_sheet(wb: Workbook, criteria: List[Dict],
                                   pairs: List[Tuple[int, int]]):
    ws = wb.create_sheet("Pairwise Survey")
    ws.sheet_properties.tabColor = BRAND_BLUE

    n = len(criteria)

    # Column widths
    ws.column_dimensions["A"].width = 6    # Question #
    ws.column_dimensions["B"].width = 28   # Criterion A
    ws.column_dimensions["C"].width = 14   # "compared to"
    ws.column_dimensions["D"].width = 28   # Criterion B
    ws.column_dimensions["E"].width = 42   # Dropdown answer
    ws.column_dimensions["F"].width = 14   # Saaty value (hidden)
    ws.column_dimensions["G"].width = 12   # Criterion A ID (hidden)
    ws.column_dimensions["H"].width = 12   # Criterion B ID (hidden)

    # Hide the formula/ID columns
    ws.column_dimensions["F"].hidden = True
    ws.column_dimensions["G"].hidden = True
    ws.column_dimensions["H"].hidden = True

    # ── Title ──
    ws.merge_cells("A1:E1")
    ws.cell(row=1, column=1, value="Pairwise Criteria Comparison").font = Font(
        name="Calibri", size=16, bold=True, color=BRAND_NAVY
    )
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:E2")
    ws.cell(row=2, column=1, value=(
        "For each pair, select which criterion is more important to your organization's AI strategy."
    )).font = HINT_FONT

    ws.merge_cells("A3:E3")
    ws.cell(row=3, column=1, value=(
        'Tip: Just pick from the dropdown -- no numbers needed. If you\'re unsure, choose "Equal importance".'
    )).font = Font(name="Calibri", size=10, italic=True, color=BRAND_BLUE)

    # ── Column Headers ──
    headers = [("#", 6), ("Criterion A", 28), ("compared to", 14),
               ("Criterion B", 28), ("Your Judgement", 42)]
    for j, (text, _) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=j, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Hidden column headers
    ws.cell(row=4, column=6, value="Saaty Value").font = SMALL_FONT
    ws.cell(row=4, column=7, value="ID_A").font = SMALL_FONT
    ws.cell(row=4, column=8, value="ID_B").font = SMALL_FONT

    # ── Data Validation (dropdown) ──
    scale_count = len(SCALE_OPTIONS)
    dv = DataValidation(
        type="list",
        formula1=f"Config!$A$2:$A${scale_count + 1}",
        allow_blank=True,
    )
    dv.prompt = "Select your judgement"
    dv.promptTitle = "Pairwise Comparison"
    dv.error = "Please select a value from the dropdown list"
    dv.errorTitle = "Invalid Selection"
    ws.add_data_validation(dv)

    # ── Data Rows ──
    data_start = 5
    for idx, (i, j) in enumerate(pairs):
        row = data_start + idx
        c_a = criteria[i]
        c_b = criteria[j]
        is_alt = idx % 2 == 1  # alternating background

        # Question number
        cell = ws.cell(row=row, column=1, value=idx + 1)
        cell.font = Font(name="Calibri", size=10, color="999999")
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if is_alt:
            cell.fill = ALT_FILL

        # Criterion A
        cell = ws.cell(row=row, column=2, value=c_a["name"])
        cell.font = Font(name="Calibri", size=11, bold=True, color=BRAND_NAVY)
        cell.alignment = LEFT_WRAP
        cell.border = THIN_BORDER
        if is_alt:
            cell.fill = ALT_FILL

        # "compared to"
        cell = ws.cell(row=row, column=3, value="compared to")
        cell.font = Font(name="Calibri", size=10, color="999999")
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if is_alt:
            cell.fill = ALT_FILL

        # Criterion B
        cell = ws.cell(row=row, column=4, value=c_b["name"])
        cell.font = Font(name="Calibri", size=11, bold=True, color=BRAND_NAVY)
        cell.alignment = LEFT_WRAP
        cell.border = THIN_BORDER
        if is_alt:
            cell.fill = ALT_FILL

        # Dropdown answer cell (user fills this)
        cell = ws.cell(row=row, column=5)
        cell.fill = INPUT_FILL
        cell.font = BODY_FONT
        cell.alignment = LEFT_WRAP
        cell.border = THIN_BORDER
        dv.add(cell)

        # Hidden: Saaty value formula (VLOOKUP from Config)
        ws.cell(row=row, column=6, value=(
            f'=IF(E{row}="","",VLOOKUP(E{row},Config!$A$2:$B${scale_count + 1},2,FALSE))'
        ))

        # Hidden: Criterion IDs (for import)
        ws.cell(row=row, column=7, value=c_a["id"])
        ws.cell(row=row, column=8, value=c_b["id"])

    data_end = data_start + len(pairs) - 1

    # ── Conditional Formatting: green tint for answered rows ──
    ws.conditional_formatting.add(
        f"A{data_start}:E{data_end}",
        FormulaRule(
            formula=[f'$E{data_start}<>""'],
            fill=GREEN_FILL,
        ),
    )

    # Freeze panes: keep headers visible
    ws.freeze_panes = "A5"


# ══════════════════════════════════════════════════════════════════════════════
#  Outcome Scoring Sheet
# ══════════════════════════════════════════════════════════════════════════════

def _create_outcome_scoring_sheet(wb: Workbook, criteria: List[Dict]):
    ws = wb.create_sheet("Outcome Scoring")
    ws.sheet_properties.tabColor = "5C6BC0"

    outcomes = [
        ("Strong Vendor\nCommitment", "Lock in with a major vendor (e.g. SAP, Microsoft, Google Cloud)"),
        ("Hybrid AI\nOperating Model", "Best-of-breed mix: vendor platforms + independent AI capabilities"),
        ("Independent AI\nControl Model", "Full in-house AI stack: maximum control and flexibility"),
    ]

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 16  # hidden score columns
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["E"].hidden = True
    ws.column_dimensions["F"].hidden = True
    ws.column_dimensions["G"].hidden = True

    # ── Title ──
    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1, value="Outcome Scoring").font = Font(
        name="Calibri", size=16, bold=True, color=BRAND_NAVY
    )
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:D2")
    ws.cell(row=2, column=1, value=(
        "For each criterion, rate how well each AI architecture outcome fits your organization."
    )).font = HINT_FONT

    ws.merge_cells("A3:D3")
    ws.cell(row=3, column=1, value=(
        'Use the dropdown to select a rating from "1 - Very Poor Fit" to "5 - Excellent Fit".'
    )).font = Font(name="Calibri", size=10, italic=True, color=BRAND_BLUE)

    # ── Headers ──
    ws.cell(row=4, column=1, value="Criterion").font = HEADER_FONT
    ws.cell(row=4, column=1).fill = HEADER_FILL
    ws.cell(row=4, column=1).border = THIN_BORDER

    for j, (name, desc) in enumerate(outcomes):
        col = j + 2
        cell = ws.cell(row=4, column=col, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Hidden headers for numeric values
    ws.cell(row=4, column=5, value="V_Score").font = SMALL_FONT
    ws.cell(row=4, column=6, value="H_Score").font = SMALL_FONT
    ws.cell(row=4, column=7, value="I_Score").font = SMALL_FONT

    # ── Data validation ──
    score_count = len(OUTCOME_SCALE)
    dv = DataValidation(
        type="list",
        formula1=f"Config!$D$2:$D${score_count + 1}",
        allow_blank=True,
    )
    dv.prompt = "Select a rating"
    dv.promptTitle = "Outcome Score"
    dv.error = "Please select from the dropdown"
    dv.errorTitle = "Invalid Score"
    ws.add_data_validation(dv)

    # ── Data Rows ──
    data_start = 5
    for i, c in enumerate(criteria):
        row = data_start + i
        is_alt = i % 2 == 1

        # Criterion name
        cell = ws.cell(row=row, column=1, value=c["name"])
        cell.font = LABEL_FONT
        cell.fill = LIGHT_FILL if not is_alt else ALT_FILL
        cell.border = THIN_BORDER
        cell.alignment = LEFT_WRAP

        for j in range(3):
            col = j + 2
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            dv.add(cell)

            # Hidden numeric value extraction formula
            # =IF(B5="","",VLOOKUP(B5,Config!$D$2:$E$6,2,FALSE))
            score_col_letter = get_column_letter(col)
            ws.cell(row=row, column=col + 3, value=(
                f'=IF({score_col_letter}{row}="","",VLOOKUP({score_col_letter}{row},Config!$D$2:$E${score_count + 1},2,FALSE))'
            ))

    data_end = data_start + len(criteria) - 1

    # Conditional formatting
    ws.conditional_formatting.add(
        f"B{data_start}:D{data_end}",
        FormulaRule(
            formula=[f'B{data_start}<>""'],
            fill=GREEN_FILL,
        ),
    )

    ws.freeze_panes = "A5"


# ══════════════════════════════════════════════════════════════════════════════
#  Matrix Sheet (Hidden — Auto-Calculated)
# ══════════════════════════════════════════════════════════════════════════════

def _create_matrix_sheet(wb: Workbook, criteria: List[Dict],
                          pairs: List[Tuple[int, int]]):
    """Auto-populated NxN pairwise matrix using formulas referencing the survey."""
    ws = wb.create_sheet("Matrix")
    ws.sheet_state = "hidden"

    n = len(criteria)

    # Build a lookup: (i, j) -> survey row number
    pair_to_survey_row = {}
    for idx, (i, j) in enumerate(pairs):
        pair_to_survey_row[(i, j)] = 5 + idx  # Survey data starts at row 5

    # ── Headers ──
    ws.cell(row=1, column=1, value="Auto-Generated Pairwise Matrix (Do Not Edit)").font = Font(
        name="Calibri", size=12, bold=True, color="FF0000"
    )

    ws.column_dimensions["A"].width = 28
    matrix_start_row = 3

    # Column headers
    ws.cell(row=matrix_start_row, column=1, value="Criterion A / B").font = LABEL_FONT
    for j, c in enumerate(criteria):
        col = j + 2
        ws.column_dimensions[get_column_letter(col)].width = 16
        cell = ws.cell(row=matrix_start_row, column=col, value=c["name"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # ── Matrix Cells ──
    for i in range(n):
        row = matrix_start_row + 1 + i
        ws.cell(row=row, column=1, value=criteria[i]["name"]).font = LABEL_FONT
        ws.cell(row=row, column=1).fill = LIGHT_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER

        for j in range(n):
            col = j + 2
            cell = ws.cell(row=row, column=col)
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            cell.number_format = '0.000'

            if i == j:
                # Diagonal = 1
                cell.value = 1
                cell.fill = DIAGONAL_FILL
                cell.font = Font(name="Calibri", size=11, color="666666")
            elif i < j:
                # Upper triangle: direct reference to survey Saaty value
                survey_row = pair_to_survey_row[(i, j)]
                cell.value = f"=IF('Pairwise Survey'!F{survey_row}=\"\",\"\", 'Pairwise Survey'!F{survey_row})"
                cell.font = BODY_FONT
            else:
                # Lower triangle: reciprocal of upper triangle
                upper_col = get_column_letter(i + 2)
                upper_row = matrix_start_row + 1 + j
                cell.value = f'=IF({upper_col}{upper_row}="","",1/{upper_col}{upper_row})'
                cell.font = Font(name="Calibri", size=11, color="888888")


# ══════════════════════════════════════════════════════════════════════════════
#  Strategic Importance Sheet
# ══════════════════════════════════════════════════════════════════════════════

def _create_strategic_importance_sheet(wb: Workbook):
    ws = wb.create_sheet("Strategic Importance")
    ws.sheet_properties.tabColor = "7986CB"

    ws.merge_cells("A1:G1")
    ws.cell(row=1, column=1, value="Strategic Importance Appendix").font = Font(
        name="Calibri", size=16, bold=True, color=BRAND_NAVY
    )
    ws.merge_cells("A2:G2")
    ws.cell(row=2, column=1, value=(
        "Rate the strategic importance of outcomes or initiatives (0 = No override, 10 = Critical imperative). "
        "Justification required for values >= 7."
    )).font = HINT_FONT

    headers = ["Item Name", "Type", "Strategic Importance (0-10)", "Justification",
               "Executive Sponsor", "Time Horizon", "Risk of Not Acting"]
    widths = [30, 15, 25, 40, 20, 15, 30]

    for j, (h, w) in enumerate(zip(headers, widths)):
        col = j + 1
        ws.column_dimensions[get_column_letter(col)].width = w
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Data validation for strategic importance (0-10)
    dv = DataValidation(type="whole", operator="between", formula1="0", formula2="10")
    dv.error = "Please enter a value between 0 and 10"
    dv.errorTitle = "Invalid Value"
    ws.add_data_validation(dv)

    # Pre-fill with 3 outcome rows
    outcomes = ["Strong Vendor Commitment", "Hybrid AI Operating Model", "Independent AI Control Model"]
    for i, outcome in enumerate(outcomes):
        row = 5 + i
        ws.cell(row=row, column=1, value=outcome).border = THIN_BORDER
        ws.cell(row=row, column=1).font = LABEL_FONT
        ws.cell(row=row, column=2, value="outcome").border = THIN_BORDER
        cell = ws.cell(row=row, column=3)
        cell.border = THIN_BORDER
        cell.fill = INPUT_FILL
        dv.add(cell)
        for j in range(4, 8):
            ws.cell(row=row, column=j).border = THIN_BORDER

    ws.freeze_panes = "A5"


# ══════════════════════════════════════════════════════════════════════════════
#  Comments Sheet
# ══════════════════════════════════════════════════════════════════════════════

def _create_comments_sheet(wb: Workbook):
    ws = wb.create_sheet("Comments")
    ws.sheet_properties.tabColor = "9FA8DA"
    ws.column_dimensions["A"].width = 80

    ws.cell(row=1, column=1, value="Comments and Observations").font = Font(
        name="Calibri", size=16, bold=True, color=BRAND_NAVY
    )
    ws.cell(row=3, column=1, value=(
        "Please enter any additional observations, concerns, or strategic context below:"
    )).font = BODY_FONT

    cell = ws.cell(row=5, column=1)
    cell.border = THIN_BORDER
    cell.fill = INPUT_FILL
    cell.alignment = LEFT_TOP
    ws.row_dimensions[5].height = 200


# ══════════════════════════════════════════════════════════════════════════════
#  Metadata Sheet
# ══════════════════════════════════════════════════════════════════════════════

def _create_metadata_sheet(wb: Workbook, project_id: str, project_name: str,
                            company_name: str, criteria: List[Dict]):
    ws = wb.create_sheet("Metadata")
    ws.sheet_properties.tabColor = "C5CAE9"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

    ws.cell(row=1, column=1, value="Metadata (Do Not Edit)").font = Font(
        name="Calibri", size=14, bold=True, color="FF0000"
    )

    metadata = [
        ("Project ID", project_id),
        ("Project Name", project_name),
        ("Company Name", company_name),
        ("Generated At", datetime.now().isoformat()),
        ("Criteria Count", str(len(criteria))),
        ("Template Version", "2.0"),
        ("Format", "question-per-row"),
        ("Application", "AlphaAlign"),
    ]

    for i, (key, value) in enumerate(metadata, 3):
        ws.cell(row=i, column=1, value=key).font = LABEL_FONT
        ws.cell(row=i, column=2, value=value).font = BODY_FONT

    # Criteria list
    row = 3 + len(metadata) + 1
    ws.cell(row=row, column=1, value="Criteria").font = LABEL_FONT
    for i, c in enumerate(criteria):
        ws.cell(row=row + 1 + i, column=1, value=c["id"]).font = SMALL_FONT
        ws.cell(row=row + 1 + i, column=2, value=c["name"]).font = BODY_FONT


# ══════════════════════════════════════════════════════════════════════════════
#  VBA Injection (Windows + Excel Only)
# ══════════════════════════════════════════════════════════════════════════════

def _try_inject_vba(xlsx_buffer: io.BytesIO) -> Optional[io.BytesIO]:
    """
    Attempt to convert .xlsx to .xlsm with VBA macros using COM automation.
    Returns the .xlsm buffer if successful, None otherwise.
    """
    try:
        import win32com.client
        import pythoncom
    except ImportError:
        return None

    try:
        pythoncom.CoInitialize()

        # Write xlsx to temp file
        tmp_dir = tempfile.mkdtemp(prefix="alphaalign_")
        xlsx_path = os.path.join(tmp_dir, "survey.xlsx")
        xlsm_path = os.path.join(tmp_dir, "survey.xlsm")

        xlsx_buffer.seek(0)
        with open(xlsx_path, "wb") as f:
            f.write(xlsx_buffer.read())

        # Open in Excel, add VBA, save as xlsm
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))

        # Add VBA module
        vba_module = wb.VBProject.VBComponents.Add(1)  # vbext_ct_StdModule
        vba_module.Name = "AlphaAlignMacros"
        vba_module.CodeModule.AddFromString(VBA_CODE)

        # Save as macro-enabled workbook
        wb.SaveAs(os.path.abspath(xlsm_path), 52)  # xlOpenXMLWorkbookMacroEnabled
        wb.Close(False)
        excel.Quit()

        # Read back the xlsm
        with open(xlsm_path, "rb") as f:
            result = io.BytesIO(f.read())
        result.seek(0)

        # Cleanup temp files
        try:
            os.remove(xlsx_path)
            os.remove(xlsm_path)
            os.rmdir(tmp_dir)
        except Exception:
            pass

        return result

    except Exception:
        return None
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
