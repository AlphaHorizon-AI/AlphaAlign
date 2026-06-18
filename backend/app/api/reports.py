"""
AlphaAlign — Report Export API endpoints.
"""

import json
import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db import get_session
from app.db.models import Project, AHPResult, Criterion, Respondent, StrategicImportance, LLMAnalysisResult

router = APIRouter()


@router.get("/projects/{project_id}/export/excel")
async def export_excel(project_id: str, session: AsyncSession = Depends(get_session)):
    """Export results as Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Get data
    project, ahp, criteria, respondents, strategic_items = await _gather_export_data(project_id, session)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    BORDER = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    # Summary sheet
    ws["A1"] = "AlphaAlign — Assessment Results"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1a237e")

    summary_data = [
        ("Company", project.company_name),
        ("Assessment", project.name),
        ("Industry", project.industry),
        ("Owner", project.owner),
        ("Status", project.status),
        ("Aggregation Method", ahp.aggregation_method if ahp else ""),
        ("Calculated At", ahp.calculated_at.isoformat() if ahp else "Not calculated"),
    ]
    for i, (key, val) in enumerate(summary_data, 3):
        ws.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val)

    if ahp:
        recommendation = json.loads(ahp.recommendation)
        outcome_scores = json.loads(ahp.outcome_scores)
        final_scores = json.loads(ahp.final_scores)
        criteria_weights = json.loads(ahp.criteria_weights)

        # Scores sheet
        ws_scores = wb.create_sheet("Scores")
        headers = ["Outcome", "Base AHP Score", "Final Score"]
        for j, h in enumerate(headers, 1):
            c = ws_scores.cell(row=1, column=j, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.border = BORDER

        labels = {"vendor": "Strong Vendor Commitment", "hybrid": "Hybrid AI Operating Model", "independent": "Independent AI Control Model"}
        for i, (key, label) in enumerate(labels.items(), 2):
            ws_scores.cell(row=i, column=1, value=label).border = BORDER
            ws_scores.cell(row=i, column=2, value=round(outcome_scores.get(key, 0), 1)).border = BORDER
            ws_scores.cell(row=i, column=3, value=round(final_scores.get(key, 0), 1)).border = BORDER

        # Criteria Weights sheet
        ws_cw = wb.create_sheet("Criteria Weights")
        for j, h in enumerate(["Criterion", "Weight (%)"], 1):
            c = ws_cw.cell(row=1, column=j, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL

        criterion_names = {c.id: c.name for c in criteria}
        sorted_weights = sorted(criteria_weights.items(), key=lambda x: x[1], reverse=True)
        for i, (cid, w) in enumerate(sorted_weights, 2):
            ws_cw.cell(row=i, column=1, value=criterion_names.get(cid, cid))
            ws_cw.cell(row=i, column=2, value=round(w * 100, 1))

    # Auto-size columns
    for ws_item in wb.worksheets:
        for col in ws_item.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws_item.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"AlphaAlign_Results_{project.company_name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/projects/{project_id}/export/pdf")
async def export_pdf(project_id: str, session: AsyncSession = Depends(get_session)):
    """Export results as PDF report."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    project, ahp, criteria, respondents, strategic_items = await _gather_export_data(project_id, session)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=25*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=24, textColor=HexColor("#1a237e"), spaceAfter=10)
    heading_style = ParagraphStyle("CustomH1", parent=styles["Heading1"], fontSize=16, textColor=HexColor("#1a237e"), spaceBefore=20)
    body_style = ParagraphStyle("CustomBody", parent=styles["Normal"], fontSize=11, spaceAfter=8)
    small_style = ParagraphStyle("Small", parent=styles["Normal"], fontSize=9, textColor=HexColor("#666666"))

    elements = []

    # Cover
    elements.append(Paragraph("AlphaAlign", title_style))
    elements.append(Paragraph("Strategic AI Choice Engine — Assessment Report", ParagraphStyle("Sub", parent=styles["Heading2"], textColor=HexColor("#3949ab"))))
    elements.append(Spacer(1, 15*mm))
    elements.append(Paragraph(f"<b>Company:</b> {project.company_name}", body_style))
    elements.append(Paragraph(f"<b>Assessment:</b> {project.name}", body_style))
    elements.append(Paragraph(f"<b>Industry:</b> {project.industry}", body_style))
    elements.append(Paragraph(f"<b>Owner:</b> {project.owner}", body_style))
    elements.append(Paragraph(f"<b>Respondents:</b> {len(respondents)}", body_style))
    elements.append(Spacer(1, 10*mm))

    if ahp:
        recommendation = json.loads(ahp.recommendation)
        outcome_scores = json.loads(ahp.outcome_scores)
        final_scores = json.loads(ahp.final_scores)
        criteria_weights = json.loads(ahp.criteria_weights)
        alignment = json.loads(ahp.alignment_data)

        # Recommendation
        elements.append(Paragraph("Recommended Position", heading_style))
        rec_label = recommendation.get("recommended_label", "")
        rec_score = recommendation.get("recommended_score", 0)
        elements.append(Paragraph(f"<b>{rec_label}</b> (Score: {rec_score}/100)", body_style))
        elements.append(Paragraph(recommendation.get("interpretation", ""), body_style))

        # Scores table
        elements.append(Paragraph("Architecture Positioning Result", heading_style))
        labels = {"vendor": "Strong Vendor Commitment", "hybrid": "Hybrid AI Operating Model", "independent": "Independent AI Control Model"}
        score_data = [["Outcome", "Base Score", "Final Score"]]
        for key, label in labels.items():
            score_data.append([label, f"{outcome_scores.get(key, 0):.1f}", f"{final_scores.get(key, 0):.1f}"])

        score_table = Table(score_data, colWidths=[200, 80, 80])
        score_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1a237e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#ffffff")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#cccccc")),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#f5f5f5"), HexColor("#ffffff")]),
        ]))
        elements.append(score_table)
        elements.append(Spacer(1, 8*mm))

        # Criteria weights
        elements.append(Paragraph("Criteria Weighting", heading_style))
        criterion_names = {c.id: c.name for c in criteria}
        weight_data = [["Criterion", "Weight (%)"]]
        for cid, w in sorted(criteria_weights.items(), key=lambda x: x[1], reverse=True):
            weight_data.append([criterion_names.get(cid, cid), f"{w*100:.1f}%"])

        weight_table = Table(weight_data, colWidths=[250, 80])
        weight_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1a237e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#ffffff")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#cccccc")),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
        ]))
        elements.append(weight_table)
        elements.append(Spacer(1, 8*mm))

        # Alignment
        elements.append(Paragraph("Leadership Alignment", heading_style))
        elements.append(Paragraph(alignment.get("summary", "No alignment data available."), body_style))

        # Roadmap
        roadmap = recommendation.get("roadmap", [])
        if roadmap:
            elements.append(Paragraph("Recommended Roadmap", heading_style))
            road_data = [["Phase", "Focus"]]
            for r in roadmap:
                road_data.append([r["phase"], r["focus"]])
            road_table = Table(road_data, colWidths=[80, 300])
            road_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1a237e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#ffffff")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#cccccc")),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            elements.append(road_table)

    elements.append(Spacer(1, 15*mm))
    elements.append(Paragraph(f"Generated by AlphaAlign — {project.updated_at.strftime('%Y-%m-%d %H:%M')}", small_style))

    doc.build(elements)
    buffer.seek(0)

    filename = f"AlphaAlign_Report_{project.company_name.replace(' ', '_')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/projects/{project_id}/export/json")
async def export_json(project_id: str, session: AsyncSession = Depends(get_session)):
    """Export full project archive as JSON."""
    project, ahp, criteria, respondents, strategic_items = await _gather_export_data(project_id, session)

    archive = {
        "application": "AlphaAlign",
        "version": "1.0.0",
        "project": {
            "id": project.id,
            "name": project.name,
            "company_name": project.company_name,
            "industry": project.industry,
            "owner": project.owner,
            "status": project.status,
        },
        "criteria": [{"id": c.id, "name": c.name, "description": c.description, "category": c.category} for c in criteria],
        "respondents": [{"id": r.id, "name": r.name, "role": r.role} for r in respondents],
    }

    if ahp:
        archive["results"] = {
            "criteria_weights": json.loads(ahp.criteria_weights),
            "outcome_scores": json.loads(ahp.outcome_scores),
            "final_scores": json.loads(ahp.final_scores),
            "alignment": json.loads(ahp.alignment_data),
            "recommendation": json.loads(ahp.recommendation),
        }

    return archive


async def _gather_export_data(project_id, session):
    result = await session.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(404, "Project not found")

    result = await session.execute(select(AHPResult).where(AHPResult.project_id == project_id).order_by(AHPResult.calculated_at.desc()))
    ahp = result.scalar_one_or_none()

    result = await session.execute(select(Criterion).where(Criterion.project_id == project_id).order_by(Criterion.sort_order))
    criteria = result.scalars().all()

    result = await session.execute(select(Respondent).where(Respondent.project_id == project_id))
    respondents = result.scalars().all()

    result = await session.execute(select(StrategicImportance).where(StrategicImportance.project_id == project_id))
    strategic_items = result.scalars().all()

    return project, ahp, criteria, respondents, strategic_items
